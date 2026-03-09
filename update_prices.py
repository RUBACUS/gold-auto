import os
import csv
import math
import logging
from datetime import datetime, timezone, timedelta

IST = timezone(timedelta(hours=5, minutes=30))

import openpyxl

from scraper import scrape_ibja_rates
from database import (
    get_latest_rate,
    save_rate,
    save_update_log,
    get_latest_diamond_rates,
    save_diamond_rates,
    save_diamond_update_log,
    get_rate_config,
    get_active_upload,
    get_upload_file_data,
    save_generated_file,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "updated_sheets")


def get_source_file():
    """Return the path to the active source CSV/xlsx file.

    Requires an active (freshly uploaded) file to be present.
    If the file is not on disk (e.g. ephemeral Railway filesystem), it is
    restored from the PostgreSQL database automatically.
    Raises ValueError if no active upload is found.
    """
    upload = get_active_upload()
    if upload:
        path = os.path.join(UPLOAD_DIR, upload["filename"])
        if os.path.isfile(path):
            return path
        # Restore from PostgreSQL (Railway ephemeral filesystem after redeploy)
        data = get_upload_file_data(upload["filename"])
        if data:
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            with open(path, "wb") as f:
                f.write(data)
            logging.info("Restored active upload '%s' from database to disk.", upload["filename"])
            return path
        raise ValueError(
            f"Active upload file not found on disk or in database: {upload['filename']}. "
            "Please upload the file again."
        )
    raise ValueError(
        "No source file uploaded. Please upload a CSV or XLSX file before running pricing."
    )


# Expected CSV header names for dynamic column detection
_EXPECTED_HEADERS = {
    "handle": "Handle",
    "opt2_value": "Option2 Value",
    "opt3_value": "Option3 Value",
    "variant_price": "Variant Price",
    "compare_at_price": "Variant Compare At Price",
}

# Metafield header patterns (case-insensitive substring match)
_METAFIELD_PATTERNS = {
    "14kt_weight": "14kt",
    "18kt_weight": "18kt",
    "9kt_weight": "9kt",
    "diamond_weight": "diamond",
    "gemstone_weight": "gemstone",
}


def _detect_csv_columns(header_row):
    """Detect column indices from CSV header row. Returns dict of name -> 0-based index.

    Falls back to hardcoded positions if headers don't match.
    """
    cols = {}
    header_lower = [h.strip().lower() for h in header_row]

    # Standard Shopify columns
    for key, expected in _EXPECTED_HEADERS.items():
        expected_lower = expected.lower()
        if expected_lower in header_lower:
            cols[key] = header_lower.index(expected_lower)

    # Weight metafield columns (match by substring in header)
    for key, pattern in _METAFIELD_PATTERNS.items():
        for i, h in enumerate(header_lower):
            if pattern in h and "weight" in h:
                cols[key] = i
                break

    return cols


# Fallback column indices (1-based, matching Shopify export) for XLSX
COL_HANDLE = 1
COL_OPT2_VALUE = 13       # Gold Quality  e.g. "14KT-Yellow"
COL_OPT3_VALUE = 16       # Diamond Quality e.g. "GH I1-I2", "GH SI"
COL_VARIANT_PRICE = 24     # Variant Price
COL_COMPARE_AT_PRICE = 25  # Variant Compare At Price
COL_14KT_WEIGHT = 50       # product metafield
COL_18KT_WEIGHT = 51       # product metafield
COL_9KT_WEIGHT = 52        # product metafield
COL_DIAMOND_WEIGHT = 55    # Diamond Total Weight (carats)
COL_GEMSTONE_WEIGHT = 60   # Gemstone Total Weight (carats)

# Fixed charges — defaults; overridden by rate_config at runtime
GOLD_MAKING_CHARGE_PER_GRAM = 2500   # ₹ per gram gold
FIXED_CHARGE = 100                    # ₹ flat per variant (HUID)
DIAMOND_LABOUR_PER_CARAT = 500       # ₹ per carat diamond (Certification)
COLORSTONE_RATE_PER_CARAT = 1500     # ₹ per carat gemstone


def ceil_safe(x):
    """math.ceil that handles exact-integer FP values correctly.

    e.g. ceil_safe(8690.0000000001) → 8690  (not 8691)
    """
    return math.ceil(round(x, 6))


def parse_weight(value):
    """Convert weight cell value (may have leading apostrophe) to float."""
    if value is None:
        return None
    s = str(value).strip().lstrip("'")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def generate_output_filename(suffix, ext=".xlsx"):
    """Create an IST-timestamped output filename."""
    now = datetime.now(IST)
    name = now.strftime(f"products_%d%b%Y_%H%M%S_IST_{suffix}{ext}")
    return os.path.join(OUTPUT_DIR, name)


def _classify_diamond_quality(opt3_value):
    """Return 'i1i2', 'si', or None for a given Option3 cell value."""
    if not opt3_value:
        return None
    val = str(opt3_value).upper()
    if "I1" in val or "I2" in val:
        return "i1i2"
    if "SI" in val:
        return "si"
    return None


def _compute_variant_price(gold_wt, gold_rate, diamond_wt, quality,
                            diamond_rates, making_charge, huid,
                            cert_rate, gem_rate, gem_wt):
    """Compute price for a single variant given rates."""
    comp_gold = ceil_safe(gold_wt * gold_rate)
    comp_gold_making = ceil_safe(gold_wt * making_charge)

    if diamond_wt > 0 and diamond_rates and quality:
        diam_rate = diamond_rates.get(quality, 0)
        comp_diamond = ceil_safe(diamond_wt * diam_rate)
        comp_diamond_labour = ceil_safe(diamond_wt * cert_rate)
    else:
        comp_diamond = 0
        comp_diamond_labour = 0

    comp_colorstone = ceil_safe(gem_wt * gem_rate) if gem_wt > 0 else 0

    return (comp_gold + comp_diamond + comp_gold_making
            + huid + comp_diamond_labour + comp_colorstone)


def _is_csv(filepath):
    return filepath.lower().endswith(".csv")


def _read_csv_rows(filepath):
    """Read CSV file and return list of rows (each row is a list of strings).

    Tries common encodings in order so Shopify exports saved as Windows-1252
    or Latin-1 are handled without error.
    """
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            rows = []
            with open(filepath, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(row)
            return rows
        except UnicodeDecodeError:
            continue
    # Last resort: replace undecodable bytes instead of crashing
    rows = []
    with open(filepath, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    return rows


def _write_csv_rows(filepath, rows):
    """Write list of rows to CSV file."""
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def update_excel_prices(gold_rates, diamond_rates=None, suffix="UNK",
                        rate_cfg=None):
    """Recalculate every 9KT, 14KT and 18KT variant price using the full formula.

    Calculates BOTH:
    - Variant Price (col 24) using the standard rate chart
    - Variant Compare At Price (col 25) using the compare-at rate chart

    Formula (round up each component, then sum):
      ceil(gold_wt × gold_rate)
    + ceil(diamond_wt × diamond_rate_per_quality)
    + ceil(gold_wt × making_charge)
    + huid_per_pc
    + ceil(diamond_wt × certification)
    + ceil(gem_wt × colorstone_rate)

    Reads from the active uploaded file, or default export.
    Supports both .xlsx and .csv input files, outputs matching format.

    gold_rates:    {"9kt": rate, "14kt": rate, "18kt": rate}
    diamond_rates: {"i1i2": rate, "si": rate} or None (uses config)
    suffix:        filename suffix (e.g. "AM", "PM", "DIAM")
    rate_cfg:      dict from get_rate_config() or None (uses defaults)

    Returns (output_path, variants_updated, products_updated).
    """
    # Load rate config (editable fields)
    if rate_cfg is None:
        rate_cfg = get_rate_config()

    # --- Standard price chart ---
    making_charge = rate_cfg.get("making_charge", GOLD_MAKING_CHARGE_PER_GRAM)
    huid = rate_cfg.get("huid_per_pc", FIXED_CHARGE)
    cert_rate = rate_cfg.get("certification", DIAMOND_LABOUR_PER_CARAT)
    gem_rate = rate_cfg.get("colorstone_rate", COLORSTONE_RATE_PER_CARAT)

    if diamond_rates is None:
        diamond_rates = {
            "i1i2": rate_cfg.get("diamond_i1i2", 0),
            "si": rate_cfg.get("diamond_si", 0),
        }

    # --- Compare-at price chart ---
    cmp_making = rate_cfg.get("cmp_making_charge", making_charge)
    cmp_huid = rate_cfg.get("cmp_huid_per_pc", huid)
    cmp_cert = rate_cfg.get("cmp_certification", cert_rate)
    cmp_gem = rate_cfg.get("cmp_colorstone_rate", gem_rate)
    cmp_diamond_rates = {
        "i1i2": rate_cfg.get("cmp_diamond_i1i2", 100000),
        "si": rate_cfg.get("cmp_diamond_si", 125000),
    }

    source_file = get_source_file()
    is_csv = _is_csv(source_file)
    rate_9 = gold_rates.get("9kt", 0)
    rate_14 = gold_rates["14kt"]
    rate_18 = gold_rates["18kt"]

    # Always output CSV for downloads
    out_ext = ".csv"

    if is_csv:
        rows = _read_csv_rows(source_file)
        if len(rows) < 2:
            raise ValueError("CSV file has no data rows")

        # Detect columns from header row and fail fast if required headers are missing.
        detected = _detect_csv_columns(rows[0])

        required_cols = [
            "handle", "opt2_value", "opt3_value", "variant_price", "compare_at_price",
            "14kt_weight", "18kt_weight", "9kt_weight", "diamond_weight", "gemstone_weight",
        ]
        missing = [k for k in required_cols if k not in detected]
        if missing:
            raise ValueError(
                "CSV header mapping failed; missing required columns: " + ", ".join(missing)
            )

        c_handle = detected["handle"]
        c_opt2 = detected["opt2_value"]
        c_opt3 = detected["opt3_value"]
        c_price = detected["variant_price"]
        c_cmp = detected["compare_at_price"]
        c_14wt = detected["14kt_weight"]
        c_18wt = detected["18kt_weight"]
        c_9wt = detected["9kt_weight"]
        c_diam = detected["diamond_weight"]
        c_gem = detected["gemstone_weight"]

        max_col = max(c_handle, c_opt2, c_opt3, c_price, c_cmp,
                      c_14wt, c_18wt, c_9wt, c_diam, c_gem)

        logging.info(f"CSV columns detected from headers: {detected}")

        # Pass 1 – build weight maps
        product_weights = {}
        product_diamond_weights = {}
        product_gemstone_weights = {}

        for ri in range(1, len(rows)):
            row = rows[ri]
            # Pad row if short
            while len(row) < max_col + 1:
                row.append("")

            handle = row[c_handle].strip()
            if not handle:
                continue

            if handle not in product_weights:
                w9 = parse_weight(row[c_9wt])
                w14 = parse_weight(row[c_14wt])
                w18 = parse_weight(row[c_18wt])
                if w9 is not None or w14 is not None or w18 is not None:
                    product_weights[handle] = {"9KT": w9, "14KT": w14, "18KT": w18}
            if handle not in product_diamond_weights:
                diam_wt = parse_weight(row[c_diam])
                if diam_wt is not None:
                    product_diamond_weights[handle] = diam_wt
            if handle not in product_gemstone_weights:
                gem_wt_val = parse_weight(row[c_gem])
                if gem_wt_val is not None:
                    product_gemstone_weights[handle] = gem_wt_val

        # Pass 2 – update prices
        variants_updated = 0
        products_touched = set()
        current_handle = None

        for ri in range(1, len(rows)):
            row = rows[ri]
            handle = row[c_handle].strip()
            if handle:
                current_handle = handle

            opt2 = row[c_opt2].strip()
            if not opt2:
                continue

            kt_str = opt2.split("-")[0].upper()
            if kt_str not in ("9KT", "14KT", "18KT"):
                continue

            weights = product_weights.get(current_handle)
            if not weights:
                continue

            gold_wt = weights.get(kt_str)
            if gold_wt is None:
                continue

            gold_rate = rate_9 if kt_str == "9KT" else rate_14 if kt_str == "14KT" else rate_18

            diamond_wt = product_diamond_weights.get(current_handle, 0) or 0
            quality = _classify_diamond_quality(row[c_opt3])
            gem_wt = product_gemstone_weights.get(current_handle, 0) or 0

            # Variant Price
            new_price = _compute_variant_price(
                gold_wt, gold_rate, diamond_wt, quality,
                diamond_rates, making_charge, huid, cert_rate, gem_rate, gem_wt)

            # Compare At Price
            cmp_price = _compute_variant_price(
                gold_wt, gold_rate, diamond_wt, quality,
                cmp_diamond_rates, cmp_making, cmp_huid, cmp_cert, cmp_gem, gem_wt)

            row[c_price] = str(new_price)
            row[c_cmp] = str(cmp_price)
            variants_updated += 1
            products_touched.add(current_handle)

        # Save CSV
        output_path = generate_output_filename(suffix, ext=out_ext)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        _write_csv_rows(output_path, rows)

        # Persist to PostgreSQL so the file survives Railway redeploys
        with open(output_path, "rb") as _f:
            save_generated_file(os.path.basename(output_path), _f.read())

    else:
        # ── XLSX path (original logic) ──
        wb = openpyxl.load_workbook(source_file)
        ws = wb.active

        # Pass 1 – build maps
        product_weights = {}
        product_diamond_weights = {}
        product_gemstone_weights = {}

        for row_idx in range(2, ws.max_row + 1):
            handle = ws.cell(row_idx, COL_HANDLE).value
            if not handle:
                continue
            if handle not in product_weights:
                w9 = parse_weight(ws.cell(row_idx, COL_9KT_WEIGHT).value)
                w14 = parse_weight(ws.cell(row_idx, COL_14KT_WEIGHT).value)
                w18 = parse_weight(ws.cell(row_idx, COL_18KT_WEIGHT).value)
                if w9 is not None or w14 is not None or w18 is not None:
                    product_weights[handle] = {"9KT": w9, "14KT": w14, "18KT": w18}
            if handle not in product_diamond_weights:
                diam_wt = parse_weight(ws.cell(row_idx, COL_DIAMOND_WEIGHT).value)
                if diam_wt is not None:
                    product_diamond_weights[handle] = diam_wt
            if handle not in product_gemstone_weights:
                gem_wt_val = parse_weight(ws.cell(row_idx, COL_GEMSTONE_WEIGHT).value)
                if gem_wt_val is not None:
                    product_gemstone_weights[handle] = gem_wt_val

        # Pass 2 – update prices
        variants_updated = 0
        products_touched = set()
        current_handle = None

        for row_idx in range(2, ws.max_row + 1):
            handle = ws.cell(row_idx, COL_HANDLE).value
            if handle:
                current_handle = handle

            opt2 = ws.cell(row_idx, COL_OPT2_VALUE).value
            if not opt2:
                continue

            kt_str = str(opt2).split("-")[0].upper()
            if kt_str not in ("9KT", "14KT", "18KT"):
                continue

            weights = product_weights.get(current_handle)
            if not weights:
                continue

            gold_wt = weights.get(kt_str)
            if gold_wt is None:
                continue

            gold_rate = rate_9 if kt_str == "9KT" else rate_14 if kt_str == "14KT" else rate_18

            diamond_wt = product_diamond_weights.get(current_handle, 0) or 0
            opt3 = ws.cell(row_idx, COL_OPT3_VALUE).value
            quality = _classify_diamond_quality(opt3)
            gem_wt = product_gemstone_weights.get(current_handle, 0) or 0

            # Variant Price
            new_price = _compute_variant_price(
                gold_wt, gold_rate, diamond_wt, quality,
                diamond_rates, making_charge, huid, cert_rate, gem_rate, gem_wt)

            # Compare At Price
            cmp_price = _compute_variant_price(
                gold_wt, gold_rate, diamond_wt, quality,
                cmp_diamond_rates, cmp_making, cmp_huid, cmp_cert, cmp_gem, gem_wt)

            ws.cell(row_idx, COL_VARIANT_PRICE).value = new_price
            ws.cell(row_idx, COL_COMPARE_AT_PRICE).value = cmp_price
            variants_updated += 1
            products_touched.add(current_handle)

        # Export as CSV (always output CSV)
        output_path = generate_output_filename(suffix, ext=out_ext)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        wb.close()

        # Persist to PostgreSQL so the file survives Railway redeploys
        with open(output_path, "rb") as _f:
            save_generated_file(os.path.basename(output_path), _f.read())

    return output_path, variants_updated, len(products_touched)


def run_update():
    """Gold update: scrape IBJA → recalculate all prices → persist.

    Uses the full formula with current gold rates + stored diamond rates.
    Returns a result dict suitable for API responses.
    """
    # 1. Scrape current IBJA rates
    ibja = scrape_ibja_rates()

    # 9KT is dynamically calculated by scraper
    rate_9kt = ibja.get("9kt", 0)

    new_rates = {
        "9kt": rate_9kt,
        "14kt": ibja["14kt"],
        "18kt": ibja["18kt"],
        "fine_gold": ibja.get("fine_gold"),
        "session": ibja["session"],
        "date": ibja["date"],
    }

    # 2. Determine old (last applied) rates
    stored = get_latest_rate()
    if stored is None:
        # First run – store current rate as baseline
        save_rate(
            new_rates["14kt"], new_rates["18kt"],
            new_rates.get("fine_gold"), new_rates["session"], new_rates["date"],
            rate_9kt=new_rates.get("9kt"),
        )
        return {
            "status": "baseline_set",
            "message": (
                "First run – current IBJA rates stored. "
                "Prices will be updated when IBJA rates change."
            ),
            "rates": new_rates,
        }

    old_rates = {"9kt": int(stored.get("rate_9kt") or 0),
                 "14kt": int(stored["rate_14kt"]),
                 "18kt": int(stored["rate_18kt"])}

    # Always proceed with CSV generation even if rates are identical

    # 3. Load rate config (all editable fields)
    rate_cfg = get_rate_config()

    # 4. Recalculate all prices using the full formula
    source = get_source_file()
    output_path, variants_updated, products_updated = update_excel_prices(
        new_rates, None, suffix=new_rates.get("session", "UNK"),
        rate_cfg=rate_cfg,
    )

    # 5. Persist new rate & log
    save_rate(
        new_rates["14kt"], new_rates["18kt"],
        new_rates.get("fine_gold"), new_rates["session"], new_rates["date"],
        rate_9kt=new_rates.get("9kt"),
    )
    save_update_log(
        old_rates["14kt"], old_rates["18kt"],
        new_rates["14kt"], new_rates["18kt"],
        os.path.basename(source),
        os.path.basename(output_path),
        variants_updated,
        products_updated,
    )

    d14 = new_rates["14kt"] - old_rates["14kt"]
    d18 = new_rates["18kt"] - old_rates["18kt"]
    d9  = new_rates.get("9kt", 0) - old_rates.get("9kt", 0)
    msg = (
        f"Prices re-exported with current rates (no rate change) for {variants_updated} variants across {products_updated} products."
        if d14 == 0 and d18 == 0 and d9 == 0
        else f"Prices updated for {variants_updated} variants across {products_updated} products."
    )

    return {
        "status": "updated",
        "message": msg,
        "input_file": os.path.basename(source),
        "output_file": os.path.basename(output_path),
        "variants_updated": variants_updated,
        "products_updated": products_updated,
        "old_rates": old_rates,
        "new_rates": new_rates,
        "delta_14kt": d14,
        "delta_18kt": d18,
        "delta_9kt": d9,
    }


# ── Diamond Price Update ─────────────────────────────────

def run_diamond_update(new_i1i2_rate, new_si_rate):
    """Manual diamond price update: recalculate all prices with new diamond rates.

    Uses the full formula with stored gold rates + new diamond rates.
    Returns a result dict suitable for API responses.
    """
    try:
        new_i1i2_rate = float(new_i1i2_rate)
        new_si_rate = float(new_si_rate)
    except (ValueError, TypeError):
        raise ValueError("Diamond rates must be valid numbers.")

    if new_i1i2_rate <= 0 or new_si_rate <= 0:
        raise ValueError("Diamond rates must be positive values.")

    new_diamond_rates = {"i1i2": new_i1i2_rate, "si": new_si_rate}

    stored = get_latest_diamond_rates()

    # First run — store as baseline, no Excel changes
    if stored is None:
        save_diamond_rates(new_i1i2_rate, new_si_rate)
        return {
            "status": "baseline_set",
            "message": (
                "Diamond rate baseline stored. "
                "Sheet will be updated the next time rates are changed."
            ),
            "rates": new_diamond_rates,
        }

    old_diamond_rates = {"i1i2": stored["rate_i1i2"], "si": stored["rate_si"]}

    if old_diamond_rates["i1i2"] == new_i1i2_rate and old_diamond_rates["si"] == new_si_rate:
        return {
            "status": "no_change",
            "message": "Diamond rates are identical to the stored baseline — no update needed.",
            "rates": new_diamond_rates,
            "old_rates": old_diamond_rates,
        }

    # Get stored gold rates (required for full formula)
    gold_stored = get_latest_rate()
    if gold_stored is None:
        raise ValueError(
            "No gold rates stored yet. Please run a gold update first "
            "so the system knows the current gold rate."
        )

    gold_rates = {"9kt": int(gold_stored.get("rate_9kt") or 0),
                  "14kt": int(gold_stored["rate_14kt"]),
                  "18kt": int(gold_stored["rate_18kt"])}

    # Load rate config and override diamond rates in it
    rate_cfg = get_rate_config()
    rate_cfg["diamond_i1i2"] = new_i1i2_rate
    rate_cfg["diamond_si"] = new_si_rate

    # Recalculate all prices
    source = get_source_file()
    output_path, variants_updated, products_updated = update_excel_prices(
        gold_rates, None, suffix="DIAM", rate_cfg=rate_cfg,
    )

    save_diamond_rates(new_i1i2_rate, new_si_rate)
    save_diamond_update_log(
        old_diamond_rates["i1i2"], old_diamond_rates["si"],
        new_i1i2_rate, new_si_rate,
        os.path.basename(source),
        os.path.basename(output_path),
        variants_updated,
        products_updated,
    )

    return {
        "status": "updated",
        "message": f"Diamond prices updated for {variants_updated} variants across {products_updated} products.",
        "input_file": os.path.basename(source),
        "output_file": os.path.basename(output_path),
        "variants_updated": variants_updated,
        "products_updated": products_updated,
        "old_rates": old_diamond_rates,
        "new_rates": new_diamond_rates,
        "delta_i1i2": new_i1i2_rate - old_diamond_rates["i1i2"],
        "delta_si": new_si_rate - old_diamond_rates["si"],
    }


if __name__ == "__main__":
    result = run_update()
    print()
    for k, v in result.items():
        print(f"  {k}: {v}")
